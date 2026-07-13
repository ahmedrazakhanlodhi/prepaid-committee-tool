"""
The 529 Network — Prepaid Committee Data Tool (internal)

Streamlit app over the consolidated prepaid tuition plan record. Explores every plan-year,
QA's completeness, and accepts new annual committee files to append (session-based, with a
downloadable new master to commit back to the repo).

Layout:
  streamlit_app.py    this app
  src/                parser
  data/               master CSV + metadata (source of truth)
  assets/             brand assets
  samples/            example exports
"""
import io
import json
import os
import re
import sys
from datetime import datetime

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, "src"))

from prepaid_parser import parse_committee_file

DATA_DIR = os.path.join(BASE_DIR, "data")
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
DATA_CSV = os.path.join(DATA_DIR, "prepaid_master.csv")
META_JSON = os.path.join(DATA_DIR, "prepaid_meta.json")
ATTRS_JSON = os.path.join(DATA_DIR, "prepaid_attrs_by_year.json")
LOGO = os.path.join(ASSETS_DIR, "logo_small.png")
LOGO_FULL = os.path.join(ASSETS_DIR, "logo.png")
LOGO_XL = os.path.join(ASSETS_DIR, "logo_excel.png")
FAVICON = os.path.join(ASSETS_DIR, "favicon.png")

# ----------------------------------------------------------------------------- config
st.set_page_config(
    page_title="Prepaid Committee Data Tool | The 529 Network",
    page_icon=FAVICON if os.path.exists(FAVICON) else "📊",
    layout="wide",
)

GREEN, DARK, MIST, STEEL = "#3A8916", "#2B650B", "#C6DDBB", "#708686"
AMBER, GREY = "#E0A100", "#EDEFEE"

METRICS = {
    "funded":                   ("Funded Status", "pct"),
    "assets_m":                 ("Assets ($M)", "money"),
    "active_accounts":          ("Active Accounts", "int"),
    "accounts_since_inception": ("Accounts Since Inception", "int"),
    "paid_out_fy_m":            ("$ Paid Out, FY ($M)", "money"),
    "paid_out_inception_m":     ("$ Paid Out Since Inception ($M)", "money"),
}
NUM_COLS = list(METRICS.keys())

st.markdown(f"""
<style>
  .stApp {{ background:#FFFFFF; }}
  h1,h2,h3 {{ color:{DARK}; }}
  div[data-testid="stMetricValue"] {{ color:{DARK}; }}
  .stTabs [data-baseweb="tab-list"] {{ gap:2px; border-bottom:1px solid {MIST}; }}
  .stTabs [data-baseweb="tab"] {{ padding:8px 16px; }}
  .stTabs [aria-selected="true"] {{ color:{DARK} !important; font-weight:600; }}
  .card {{ border:1px solid {MIST}; border-radius:10px; padding:14px 16px; background:#FAFCF8; }}
  .brandbar {{ border-bottom:2px solid {GREEN}; padding-bottom:10px; margin-bottom:6px; }}
  .subtle {{ color:{STEEL}; font-size:0.86rem; }}
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------------------------------- data
@st.cache_data
def load_meta():
    m = json.load(open(META_JSON))
    return m["attr"], m["reg"]

@st.cache_data
def load_master():
    df = pd.read_csv(DATA_CSV)
    for c in NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["reporting_year"] = df["reporting_year"].astype(int)
    return df

ATTR, REG = load_meta()

def enrich(records, year, source_label):
    """Turn parser records (plan_key + metrics) into full master rows via the registry."""
    rows = []
    for r in records:
        k = r["plan_key"]
        state, name, est, status, sort = REG.get(k, [k, k, "", "", 99])
        rows.append({
            "plan_key": k, "state": state, "plan_name": name, "established": est,
            "status": status, "reporting_year": year, "as_of": r.get("as_of", ""),
            "funded": r.get("funded"), "assets_m": r.get("assets_m"),
            "active_accounts": r.get("active_accounts"),
            "accounts_since_inception": r.get("accounts_since_inception"),
            "paid_out_fy_m": r.get("paid_out_fy_m"),
            "paid_out_inception_m": r.get("paid_out_inception_m"),
            "source": source_label, "note": r.get("note", ""), "sort": sort,
        })
    return pd.DataFrame(rows)

if "df" not in st.session_state:
    st.session_state.df = load_master().copy()

@st.cache_data
def load_attrs_by_year():
    try:
        return json.load(open(ATTRS_JSON))
    except Exception:
        return {}

if "attrs_by_year" not in st.session_state:
    st.session_state.attrs_by_year = {k: v for k, v in load_attrs_by_year().items()}

def df():
    return st.session_state.df

# ----------------------------------------------------------------------------- helpers
def plan_order(keys):
    return sorted(keys, key=lambda k: REG.get(k, ["", "", "", "", 99])[4])

def name_of(k):
    return REG.get(k, [k, k])[1]

def fmt(val, kind):
    if pd.isna(val): return "—"
    if kind == "pct":   return f"{val*100:,.1f}%"
    if kind == "money": return f"${val:,.1f}M"
    if kind == "int":   return f"{int(val):,}"
    return str(val)

def latest_per_plan(d, col):
    """Most recent non-null value of col per plan_key."""
    out = {}
    for k, g in d.dropna(subset=[col]).groupby("plan_key"):
        g = g.sort_values("reporting_year")
        out[k] = (g.iloc[-1][col], int(g.iloc[-1]["reporting_year"]))
    return out

def build_workbook(d):
    """Rebuild a consolidated workbook (README + Master + wide matrices + attributes)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as L
    wb = openpyxl.Workbook()
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    bf = Font(name="Arial", size=10)
    fill = PatternFill("solid", fgColor=GREEN.lstrip("#"))
    grp = PatternFill("solid", fgColor=MIST.lstrip("#"))
    thin = Side(style="thin", color="D9D9D9"); bd = Border(thin, thin, thin, thin)
    years = sorted(d["reporting_year"].unique())

    # README
    ws = wb.active; ws.title = "README"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 110; ws.column_dimensions["A"].width = 3
    ws.row_dimensions[1].height = 34
    _embed_logo(ws, "B1", height=32)
    ws.cell(2, 2, "The 529 Network — Prepaid Plan Data (consolidated)").font = Font(name="Arial", bold=True, size=14, color=DARK.lstrip("#"))
    ws.cell(3, 2, f"Generated by the Prepaid Committee Data Tool on {datetime.now():%Y-%m-%d}. One row per plan per reporting period.").font = Font(name="Arial", italic=True, size=9, color="555555")
    ws.cell(5, 2, f"Reporting years present: {', '.join(str(y) for y in years)}").font = bf
    ws.cell(6, 2, f"Plans: {d['plan_key'].nunique()}   |   Rows: {len(d)}").font = bf
    ws.cell(7, 2, "Funded status is a fraction shown as a percent (1.57 = 157%). Assets and payouts are in $ millions. Accounts are counts.").font = bf

    # Master (Long)
    ws = wb.create_sheet("Master (Long)")
    cols = ["state","plan_name","status","reporting_year","as_of","funded","assets_m",
            "active_accounts","accounts_since_inception","paid_out_fy_m","paid_out_inception_m","source","note"]
    heads = ["State","Plan Name","Status","Year","As-Of","Funded Status","Assets ($M)",
             "Active Accounts","Accounts Since Inception","$ Paid Out FY ($M)","$ Paid Out Since Incep. ($M)","Source","Notes"]
    ws.append(heads)
    for c in range(1, len(heads)+1):
        ws.cell(1, c).font = hf; ws.cell(1, c).fill = fill
        ws.cell(1, c).alignment = Alignment(horizontal="center", wrap_text=True)
    dd = d.sort_values(["sort","reporting_year"])
    for _, r in dd.iterrows():
        ws.append([r.get(c) if not pd.isna(r.get(c)) else None for c in cols])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row: c.font = bf; c.border = bd
        row[5].number_format = "0.0%"
        for i in (6, 9, 10): row[i].number_format = "#,##0.0"
        for i in (7, 8): row[i].number_format = "#,##0"
    ws.freeze_panes = "A2"
    for i, w in enumerate([7,44,14,7,12,13,12,14,15,15,16,22,38], 1):
        ws.column_dimensions[L(i)].width = w

    # wide matrices
    def wide(title, metric, numfmt):
        ws = wb.create_sheet(title)
        head = ["State","Plan Name","Status"] + [str(y) for y in years]
        ws.append(head)
        for c in range(1, len(head)+1):
            ws.cell(1, c).font = hf; ws.cell(1, c).fill = fill
            ws.cell(1, c).alignment = Alignment(horizontal="center")
        look = {(r["plan_key"], int(r["reporting_year"])): r[metric] for _, r in d.iterrows()}
        cur, rn = None, 2
        for k in plan_order(d["plan_key"].unique()):
            state, name, est, status, _ = REG.get(k, [k, k, "", "", 99])
            oc = "Open" if status == "Open" else "Closed"
            if cur != oc:
                ws.append([f"{oc.upper()} PLANS"] + [""]*(len(head)-1))
                for c in range(1, len(head)+1):
                    ws.cell(rn, c).fill = grp; ws.cell(rn, c).font = Font(name="Arial", bold=True)
                cur = oc; rn += 1
            vals = [look.get((k, y)) for y in years]
            vals = [None if (v is None or (isinstance(v, float) and pd.isna(v))) else v for v in vals]
            ws.append([state, name, status] + vals); rn += 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=len(head)):
            for c in row:
                if c.value is not None and not isinstance(c.value, str):
                    c.number_format = numfmt; c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "D2"
        ws.column_dimensions["A"].width = 7; ws.column_dimensions["B"].width = 44; ws.column_dimensions["C"].width = 14
        for i in range(4, len(head)+1): ws.column_dimensions[L(i)].width = 9
    wide("Funded Status", "funded", "0.0%")
    wide("Assets ($M)", "assets_m", "#,##0")
    wide("Active Accounts", "active_accounts", "#,##0")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# --------------------------------------------------------- committee-format export
def _embed_logo(ws, anchor="A1", height=34):
    """Drop the brand logo into a worksheet. Silently skips if the asset is absent."""
    try:
        from openpyxl.drawing.image import Image as XLImage
        path = LOGO_XL if os.path.exists(LOGO_XL) else LOGO_FULL
        if not os.path.exists(path):
            return
        img = XLImage(path)
        ratio = img.width / img.height
        img.height = height
        img.width = int(height * ratio)
        ws.add_image(img, anchor)
    except Exception:
        pass

OPEN_DISP = ["AK", "FL", "MA", "MD", "MI", "MS", "NV", "PA", "TX-II", "VA", "WA", "US"]
CLOSED_DISP = ["AL", "IL", "KY", "OH", "SC", "TX-I", "TN", "WV", "CO"]
COMMITTEE_HEADS = ["State", "Plan Name", "Year Established", "Full Faith and Credit",
                   "State Tax Exemption", "Benefit Structure", "Minimum Benefits / Alternative",
                   "Tuition Growth", "Investment Return", "Funded Status", "As of Date",
                   "Active Accounts", "Current Assets", "Accounts Since Inception",
                   "$ Paid Out in Most Recent FY", "$ Paid Out Since Inception", "Additional Notes"]

def _money_str(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ""
    v = float(v)
    if v >= 1000: return f"${v/1000:g}B"     # 1330 -> $1.33B, not $1.3B
    return f"${v:g}M"

def _pct_str(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ""
    return f"{v*100:,.1f}%"

def _int_str(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return ""
    return f"{int(v):,}"

def _note(x):
    if x is None or (isinstance(x, float) and pd.isna(x)): return ""
    return str(x)

def _fmt_assumption(v):
    """Tidy an assumption cell: clean '%%', turn plain decimals (0.03) into percents (3%),
    leave ranges and descriptive text as entered."""
    s = str(v or "").strip()
    if not s or s.lower() in ("nan", "none"): return ""
    s = s.replace("%%", "%")
    if re.fullmatch(r"\d*\.?\d+", s):
        try:
            f = float(s)
            if f <= 1.0:
                return f"{f*100:g}%"
        except ValueError:
            pass
    return s

def _attr_for(dk, ya, year=None, all_years=None):
    """Descriptive columns for a display key. Uses the year's own values, then falls back to
    the most recent earlier year that reported them (a plan may skip a year, but its tax
    treatment and benefit structure do not change), then to the static registry."""
    akey = "MI-II" if dk == "MI" else dk
    a = ya.get(akey) or (ya.get("MI-I") if dk == "MI" else None) or {}
    static = ATTR.get("MI" if dk in ("MI", "MI-I", "MI-II") else dk, {})
    regkey = "MI-II" if dk == "MI" else dk
    est = REG.get(regkey, ["", "", "", "", 0])[2]
    est = str(est).replace(".0", "") if est else (str(a.get("established", "")).replace(".0", ""))

    def clean(v):
        v = str(v or "").strip()
        return "" if ("unavailable" in v.lower() or v.lower() in ("nan", "none")) else v

    def pick(field):
        v = clean(a.get(field))
        if v:
            return v
        if all_years and year is not None:          # carry forward from the last year that reported it
            for y in sorted((int(k) for k in all_years if str(k).isdigit() and int(k) < year), reverse=True):
                prev = (all_years.get(str(y), {}) or {}).get(akey, {}) or {}
                v = clean(prev.get(field))
                if v:
                    return v
        return clean(static.get(field))

    return {
        "established": est,
        "full_faith": pick("full_faith"), "tax": pick("tax"), "benefit": pick("benefit"),
        "min_benefits": pick("min_benefits"),
        "tuition_growth": _fmt_assumption(pick("tuition_growth")),
        "inv_return": _fmt_assumption(pick("inv_return")),
    }

def _combine_mi(mi1, mi2, col, formatter):
    parts = []
    if mi1 is not None and pd.notna(mi1.get(col)): parts.append(f"MET I: {formatter(mi1[col])}")
    if mi2 is not None and pd.notna(mi2.get(col)): parts.append(f"MET II: {formatter(mi2[col])}")
    return "  ".join(parts)

def committee_rows(d, year, ya, all_years=None):
    """Return (open_rows, closed_rows); each row is a dict keyed by COMMITTEE_HEADS plus _funded_num."""
    dy = d[d["reporting_year"] == year]
    by_key = {r["plan_key"]: r for _, r in dy.iterrows()}

    def build(dk):
        at = _attr_for(dk, ya, year, all_years)
        base = {"State": "U.S." if dk == "US" else REG.get(dk if dk != "MI" else "MI-II", [dk])[0],
                "Plan Name": "Michigan Education Trust (MET)" if dk == "MI" else REG.get(dk, ["", dk])[1],
                "Year Established": at["established"], "Full Faith and Credit": at["full_faith"],
                "State Tax Exemption": at["tax"], "Benefit Structure": at["benefit"],
                "Minimum Benefits / Alternative": at["min_benefits"],
                "Tuition Growth": at["tuition_growth"], "Investment Return": at["inv_return"],
                "_funded_num": None}
        if dk == "MI":
            mi1 = by_key.get("MI-I"); mi2 = by_key.get("MI-II")
            if mi1 is None and mi2 is None: return None
            primary = mi2 if mi2 is not None else mi1
            base["Funded Status"] = _combine_mi(mi1, mi2, "funded", lambda v: _pct_str(v))
            base["As of Date"] = primary.get("as_of", "")
            base["Active Accounts"] = _combine_mi(mi1, mi2, "active_accounts", _int_str)
            base["Current Assets"] = _combine_mi(mi1, mi2, "assets_m", _money_str)
            base["Accounts Since Inception"] = _combine_mi(mi1, mi2, "accounts_since_inception", _int_str)
            base["$ Paid Out in Most Recent FY"] = _combine_mi(mi1, mi2, "paid_out_fy_m", _money_str)
            base["$ Paid Out Since Inception"] = _combine_mi(mi1, mi2, "paid_out_inception_m", _money_str)
            nt = _note(primary.get("note"))
            nt = re.sub(r"MET I+ \([^)]*\)", "", nt)
            base["Additional Notes"] = nt.strip(" ;")
            return base
        r = by_key.get(dk)
        if r is None: return None
        note = _note(r.get("note"))
        funded = r.get("funded")
        if ("Legacy tier" in note) and pd.notna(funded):
            m = re.search(r"Legacy tier ([\d\.]+)%", note)
            base["Funded Status"] = f"Horizon: {_pct_str(funded)}" + (f"  Legacy: {m.group(1)}%" if m else "")
        else:
            base["Funded Status"] = _pct_str(funded) if pd.notna(funded) else ""
            base["_funded_num"] = funded if pd.notna(funded) else None
        base["As of Date"] = r.get("as_of", "")
        base["Active Accounts"] = _int_str(r.get("active_accounts"))
        base["Current Assets"] = _money_str(r.get("assets_m"))
        base["Accounts Since Inception"] = _int_str(r.get("accounts_since_inception"))
        base["$ Paid Out in Most Recent FY"] = _money_str(r.get("paid_out_fy_m"))
        base["$ Paid Out Since Inception"] = _money_str(r.get("paid_out_inception_m"))
        base["Additional Notes"] = re.sub(r"Legacy tier [\d\.]+%;?\s*", "", note).strip(" ;")
        return base

    open_rows = [b for b in (build(dk) for dk in OPEN_DISP) if b]
    closed_rows = [b for b in (build(dk) for dk in CLOSED_DISP) if b]
    return open_rows, closed_rows

def write_committee_sheet(ws, d, year, ya, all_years=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as L
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    bf = Font(name="Arial", size=9)
    fill = PatternFill("solid", fgColor=GREEN.lstrip("#"))
    grp = PatternFill("solid", fgColor=MIST.lstrip("#"))
    thin = Side(style="thin", color="D9D9D9"); bd = Border(thin, thin, thin, thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(vertical="top", wrap_text=True)

    ws.row_dimensions[1].height = 30
    _embed_logo(ws, "A1", height=30)
    ws.cell(1, 3, f"Prepaid Plan Statistics — as reported for {year}").font = Font(name="Arial", bold=True, size=12, color=DARK.lstrip("#"))
    ws.cell(2, 3, "The 529 Network · compiled from state-submitted committee data").font = Font(name="Arial", italic=True, size=8, color=STEEL.lstrip("#"))
    # header rows 3 (super) and 4 (sub)
    n = len(COMMITTEE_HEADS)
    for c in range(1, n + 1):
        ws.cell(4, c).fill = fill; ws.cell(4, c).border = bd
    labels3 = {1:"State",2:"Plan Name",3:"Year Established",4:"Full Faith and Credit",
               5:"State Tax Exemption",6:"Benefit Structure",7:"Minimum Benefits / Alternative",
               8:"Actuarial Assumptions",10:"Funded Status",12:"Active Accounts",13:"Current Assets",
               14:"Accounts Since Inception",15:"$ Paid Out in Most Recent FY",
               16:"$ Paid Out Since Inception",17:"Additional Notes"}
    for c, t in labels3.items():
        cell = ws.cell(3, c, t); cell.font = hf; cell.fill = fill; cell.alignment = center; cell.border = bd
    ws.merge_cells("H3:I3"); ws.merge_cells("J3:K3")
    for c in (1,2,3,4,5,6,7,12,13,14,15,16,17):
        ws.merge_cells(start_row=3, start_column=c, end_row=4, end_column=c)
    for c, t in {8:"Tuition Growth",9:"Investment Return",10:"Funded Status",11:"As of Date"}.items():
        cell = ws.cell(4, c, t); cell.font = hf; cell.fill = fill; cell.alignment = center; cell.border = bd

    open_rows, closed_rows = committee_rows(d, year, ya, all_years)
    r = 5
    for title, rows in [("OPEN PLANS", open_rows), ("CLOSED PLANS", closed_rows)]:
        if not rows: continue
        ws.cell(r, 1, title).font = Font(name="Arial", bold=True, size=9)
        for c in range(1, n + 1):
            ws.cell(r, c).fill = grp; ws.cell(r, c).border = bd
        r += 1
        for row in rows:
            for c, h in enumerate(COMMITTEE_HEADS, 1):
                if h == "Funded Status" and row.get("_funded_num") is not None:
                    cell = ws.cell(r, c, row["_funded_num"]); cell.number_format = "0.0%"
                else:
                    cell = ws.cell(r, c, row.get(h, ""))
                cell.font = bf; cell.border = bd
                cell.alignment = center if c >= 8 else left
            r += 1
    widths = [7, 34, 10, 12, 20, 16, 18, 16, 16, 16, 12, 16, 18, 16, 16, 16, 34]
    for i, w in enumerate(widths, 1): ws.column_dimensions[L(i)].width = w
    ws.freeze_panes = "A5"

def committee_workbook(d, years, attrs_by_year):
    import openpyxl
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for y in years:
        ws = wb.create_sheet(str(y))
        write_committee_sheet(ws, d, y, attrs_by_year.get(str(y), {}), attrs_by_year)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def template_bytes():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 3, "Prepaid Plan Statistics — Committee Template")
    heads = ["State","Plan Name","Year Established","Full Faith and Credit","State Tax Exemption",
             "Benefit Structure","Minimum Benefits","Tuition Growth","Investment Return","Funded Status",
             "As of Date","Active Accounts","Current Assets","Accounts Since Inception",
             "$ Paid Out in Most Recent FY","$ Paid Out Since Inception","Additional Notes"]
    for c, h in enumerate(heads, 1):
        ws.cell(3, c, h).font = Font(bold=True, color="FFFFFF")
        ws.cell(3, c).fill = PatternFill("solid", fgColor=GREEN.lstrip("#"))
        ws.cell(3, c).alignment = Alignment(wrap_text=True, horizontal="center")
    ws.cell(5, 1, "OPEN PLANS")
    order = plan_order([k for k in REG if REG[k][3] == "Open"])
    r = 6
    for k in order:
        st_, name, est, status, _ = REG[k]
        if k == "MI-II": continue
        code = "U.S." if k == "US" else st_
        ws.append([code, name, est] + [""]*14) if False else None
        for c, v in enumerate([code, name, est], 1): ws.cell(r, c, v)
        r += 1
    ws.cell(r, 1, "CLOSED PLANS"); r += 1
    for k in plan_order([k for k in REG if REG[k][3] != "Open"]):
        st_, name, est, status, _ = REG[k]
        if k == "MI-I": name = "Michigan Education Trust (report MET I and MET II in each cell)"
        code = "U.S." if k == "US" else st_
        for c, v in enumerate([code, name, est], 1): ws.cell(r, c, v)
        r += 1
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ----------------------------------------------------------------------------- header
hc = st.columns([1, 3])
with hc[0]:
    if os.path.exists(LOGO):
        st.image(LOGO)
with hc[1]:
    st.markdown(
        f"<div class='brandbar'><h1 style='margin:0;padding:0;font-size:1.9rem;'>Prepaid Committee Data Tool</h1>"
        f"<div class='subtle'>Prepaid tuition plan statistics, {int(df()['reporting_year'].min())}–{int(df()['reporting_year'].max())} · internal</div></div>",
        unsafe_allow_html=True)

tabs = st.tabs(["Overview", "Plan Profile", "Compare", "Trends", "Committee Era (2022+)",
                "Data Quality", "Upload & Append", "Exports"])

# =========================================================== OVERVIEW
with tabs[0]:
    d = df()
    years = sorted(d["reporting_year"].unique())
    latest = max(years)
    dl = d[d["reporting_year"] == latest]
    tot_assets = dl["assets_m"].sum(skipna=True)
    tot_acct = dl["active_accounts"].sum(skipna=True)
    n_report = dl["assets_m"].notna().sum()
    n_open = sum(1 for k in d["plan_key"].unique() if REG.get(k, ["","","","",0])[3] == "Open")
    n_closed = d["plan_key"].nunique() - n_open

    c = st.columns(5)
    c[0].metric("Plans tracked", d["plan_key"].nunique())
    c[1].metric("Open / Closed", f"{n_open} / {n_closed}")
    c[2].metric(f"Assets, {latest}", f"${tot_assets/1000:,.1f}B",
                help=f"Sum across the {n_report} plans that reported assets in {latest}. "
                     "Plans whose last report predates this year are excluded rather than carried forward.")
    c[3].metric(f"Active accounts, {latest}", f"{tot_acct:,.0f}",
                help=f"Sum across plans reporting in {latest}.")
    c[4].metric("Reporting periods", f"{len(years)}  ({min(years)}–{max(years)})")

    st.markdown("#### Reporting depth")
    st.caption("Each cell shows how many of the six tracked metrics a plan reported that year "
               "(funded status, assets, active accounts, accounts since inception, and the two payout figures). "
               "Darker means a fuller record. Blank means the year was not collected for that plan. "
               "Hover for the detail; an ⚠ marks a figure carried forward or reported as of an earlier date.")

    keys = plan_order(d["plan_key"].unique())
    zmap, hover, marks = [], [], []
    for k in keys:
        zrow, hrow, mrow = [], [], []
        for y in years:
            sub = d[(d["plan_key"] == k) & (d["reporting_year"] == y)]
            if sub.empty:
                zrow.append(None); hrow.append(f"<b>{name_of(k)}</b> · {y}<br>Not collected"); mrow.append("")
                continue
            r0 = sub.iloc[0]
            present = [METRICS[m][0] for m in NUM_COLS if pd.notna(r0[m])]
            note = str(r0["note"]).lower()
            asof = str(r0["as_of"])
            m = re.search(r"(20\d\d)", asof)
            lagged = bool(m) and int(m.group(1)) < y
            flagged = ("unavailable" in note) or ("carried" in note) or lagged
            zrow.append(len(present))
            bits = [f"<b>{name_of(k)}</b> · {y}"]
            if not present:
                bits.append("Row present, no figures reported")
            else:
                bits.append(f"{len(present)} of {len(NUM_COLS)} metrics reported")
                bits.append("· " + "<br>· ".join(present))
            if asof and asof.lower() not in ("nan", ""):
                bits.append(f"As of {asof}")
            if flagged:
                bits.append("⚠ carried forward or lagged as-of")
            hrow.append("<br>".join(bits))
            mrow.append("⚠" if flagged else "")
        zmap.append(zrow); hover.append(hrow); marks.append(mrow)

    fig = go.Figure(go.Heatmap(
        z=zmap, x=[str(y) for y in years], y=[name_of(k) for k in keys],
        text=marks, texttemplate="%{text}", textfont=dict(size=11, color="#7A4B00"),
        customdata=hover, hovertemplate="%{customdata}<extra></extra>",
        colorscale=[[0.00, "#E4E4E2"], [0.16, "#E4E4E2"],      # 0 metrics: grey
                    [0.17, "#EAF3E4"], [0.50, MIST],
                    [0.75, "#7BB661"], [1.00, DARK]],
        zmin=0, zmax=6, xgap=2, ygap=2,
        colorbar=dict(title=dict(text="Metrics<br>reported", font=dict(size=11)),
                      tickvals=[0, 1, 2, 3, 4, 5, 6], thickness=12, len=0.55, outlinewidth=0),
        hoverongaps=False))
    fig.update_layout(height=580, margin=dict(l=10, r=10, t=10, b=10),
                      yaxis=dict(autorange="reversed"), plot_bgcolor="white", paper_bgcolor="white")
    st.plotly_chart(fig, width='stretch')

    # era summary: average depth by year, which is the real story
    depth = (d.set_index(["plan_key", "reporting_year"])[NUM_COLS].notna().sum(axis=1)
              .groupby(level=1).mean().round(1))
    ec = st.columns(3)
    ec[0].metric("Avg metrics/plan, 2005–2015", f"{depth.loc[[y for y in depth.index if y <= 2015]].mean():.1f} of 6")
    mid = [y for y in depth.index if 2016 <= y <= 2021]
    ec[1].metric("Avg metrics/plan, 2018–2021", f"{depth.loc[mid].mean():.1f} of 6" if mid else "—")
    era = [y for y in depth.index if y >= 2022]
    ec[2].metric("Avg metrics/plan, 2022+", f"{depth.loc[era].mean():.1f} of 6" if era else "—")

    missing_years = [y for y in range(min(years), max(years)+1) if y not in years]
    if missing_years:
        st.info("Years not yet collected anywhere in the record: "
                + ", ".join(str(y) for y in missing_years)
                + ".  Add them any time from the Upload & Append tab.")

# =========================================================== PLAN PROFILE
with tabs[1]:
    d = df()
    keys = plan_order(d["plan_key"].unique())
    k = st.selectbox("Plan", keys, format_func=name_of)
    sub = d[d["plan_key"] == k].sort_values("reporting_year")
    state, name, est, status, _ = REG.get(k, [k, k, "", "", 0])
    ak = "MI" if k in ("MI-I", "MI-II") else k
    at = ATTR.get(ak, {})

    st.markdown(f"### {name}")
    cc = st.columns(4)
    cc[0].markdown(f"<div class='card'><b>State</b><br>{state}</div>", unsafe_allow_html=True)
    cc[1].markdown(f"<div class='card'><b>Established</b><br>{est}</div>", unsafe_allow_html=True)
    cc[2].markdown(f"<div class='card'><b>Status</b><br>{status}</div>", unsafe_allow_html=True)
    cc[3].markdown(f"<div class='card'><b>Full faith & credit</b><br>{at.get('full_faith','—') or '—'}</div>", unsafe_allow_html=True)
    cc2 = st.columns(4)
    cc2[0].markdown(f"<div class='card'><b>State tax exemption</b><br>{at.get('tax','—') or '—'}</div>", unsafe_allow_html=True)
    cc2[1].markdown(f"<div class='card'><b>Benefit structure</b><br>{at.get('benefit','—') or '—'}</div>", unsafe_allow_html=True)
    cc2[2].markdown(f"<div class='card'><b>Tuition growth (2025)</b><br>{at.get('tuition_growth','—') or '—'}</div>", unsafe_allow_html=True)
    cc2[3].markdown(f"<div class='card'><b>Investment return (2025)</b><br>{at.get('inv_return','—') or '—'}</div>", unsafe_allow_html=True)

    st.markdown("#### History")
    mcols = st.columns(3)
    for i, m in enumerate(["funded", "assets_m", "active_accounts"]):
        g = sub.dropna(subset=[m])
        fig = go.Figure()
        if not g.empty:
            fig.add_trace(go.Scatter(x=g["reporting_year"], y=g[m], mode="lines+markers",
                          line=dict(color=GREEN, width=3), marker=dict(size=7, color=DARK)))
        yfmt = ".0%" if m == "funded" else ",.0f"
        fig.update_layout(title=METRICS[m][0], height=260, margin=dict(l=8, r=8, t=34, b=8),
                          plot_bgcolor="white", yaxis=dict(tickformat=yfmt), showlegend=False)
        mcols[i].plotly_chart(fig, width='stretch')

    show = sub[["reporting_year","as_of","funded","assets_m","active_accounts",
                "accounts_since_inception","paid_out_fy_m","paid_out_inception_m","note"]].copy()
    show.columns = ["Year","As-Of","Funded","Assets ($M)","Active Accts","Accts Since Incep.",
                    "Paid FY ($M)","Paid Since Incep. ($M)","Notes"]
    show["Funded"] = show["Funded"].apply(lambda v: fmt(v, "pct"))
    st.dataframe(show, width='stretch', hide_index=True)

# =========================================================== COMPARE
with tabs[2]:
    d = df()
    keys = plan_order(d["plan_key"].unique())
    cc = st.columns([2, 1, 1])
    sel = cc[0].multiselect("Plans", keys, default=keys[:6], format_func=name_of)
    metric = cc[1].selectbox("Metric", NUM_COLS, format_func=lambda m: METRICS[m][0])
    years = sorted(d["reporting_year"].unique())
    yr = cc[2].selectbox("Year", ["Latest"] + [str(y) for y in reversed(years)])

    rows = []
    for k in sel:
        g = d[d["plan_key"] == k].dropna(subset=[metric])
        if g.empty: continue
        if yr == "Latest":
            g = g.sort_values("reporting_year"); r = g.iloc[-1]
        else:
            gg = g[g["reporting_year"] == int(yr)]
            if gg.empty: continue
            r = gg.iloc[0]
        rows.append((name_of(k), r[metric], int(r["reporting_year"])))
    if rows:
        rows.sort(key=lambda x: x[1], reverse=True)
        kind = METRICS[metric][1]
        mixed = len({r[2] for r in rows}) > 1
        labels = [f"{fmt(r[1], kind)}  ({r[2]})" if mixed else fmt(r[1], kind) for r in rows]
        fig = go.Figure(go.Bar(
            x=[r[1] for r in rows], y=[r[0] for r in rows], orientation="h",
            marker_color=GREEN, text=labels, textposition="auto"))
        if mixed:
            st.caption("These plans last reported in different years. The source year is shown on each bar.")
        fig.update_layout(height=40*len(rows)+80, margin=dict(l=8, r=8, t=10, b=8),
                          plot_bgcolor="white", yaxis=dict(autorange="reversed"),
                          xaxis=dict(tickformat=".0%" if kind == "pct" else ",.0f"))
        st.plotly_chart(fig, width='stretch')
        tbl = pd.DataFrame({"Plan": [r[0] for r in rows],
                            METRICS[metric][0]: [fmt(r[1], kind) for r in rows],
                            "Year": [r[2] for r in rows]})
        st.dataframe(tbl, width='stretch', hide_index=True)
    else:
        st.warning("No data for that selection.")

# =========================================================== TRENDS
with tabs[3]:
    d = df()
    keys = plan_order(d["plan_key"].unique())
    cc = st.columns([1, 2])
    metric = cc[0].selectbox("Metric ", NUM_COLS, format_func=lambda m: METRICS[m][0], key="tm")
    default = [k for k in keys if REG.get(k, ["","","","",0])[3] == "Open"][:6]
    sel = cc[1].multiselect("Plans ", keys, default=default, format_func=name_of, key="ts")
    palette = [GREEN, DARK, STEEL, "#8FbF6f", "#9C7A00", "#4C7FB0", "#B0574C", "#6B6B6B",
               "#2E8B57", "#C0873B", "#5B8C5A", "#8A6D3B"]
    fig = go.Figure()
    for i, k in enumerate(sel):
        g = d[d["plan_key"] == k].dropna(subset=[metric]).sort_values("reporting_year")
        if g.empty: continue
        fig.add_trace(go.Scatter(x=g["reporting_year"], y=g[metric], mode="lines+markers",
                      name=name_of(k), line=dict(width=2.5, color=palette[i % len(palette)])))
    kind = METRICS[metric][1]
    fig.update_layout(height=520, margin=dict(l=8, r=8, t=10, b=8), plot_bgcolor="white",
                      yaxis=dict(tickformat=".0%" if kind == "pct" else ",.0f"),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02))
    st.plotly_chart(fig, width='stretch')

# =========================================================== COMMITTEE ERA (2022+)
with tabs[4]:
    d = df()
    aby = st.session_state.attrs_by_year
    era_years = [y for y in sorted(d["reporting_year"].unique()) if y >= 2022]
    st.caption("The committee collection from 2022 on carries the full detail set — actuarial "
               "assumptions, benefit structure, tax treatment, accounts-since-inception, and payouts — "
               "that the earlier record does not. Everything in this tab is drawn from those complete years.")

    if not era_years:
        st.info("No 2022-or-later years in the record yet. Add one from Upload & Append.")
    else:
        view = st.radio("View", ["Single-year detail", "Assumptions across years", "Metric matrix"],
                        horizontal=True)

        if view == "Single-year detail":
            yr = st.selectbox("Year", list(reversed(era_years)))
            op_rows, cl_rows = committee_rows(d, yr, aby.get(str(yr), {}), aby)
            det = pd.DataFrame(
                [{"Section": "Open", **{k: v for k, v in r.items() if k != "_funded_num"}} for r in op_rows] +
                [{"Section": "Closed", **{k: v for k, v in r.items() if k != "_funded_num"}} for r in cl_rows])
            st.dataframe(det, width='stretch', hide_index=True, height=560)
            st.caption("Same content as the committee-format export for this year. Download it from the Exports tab.")

        elif view == "Assumptions across years":
            st.markdown("These columns exist only for the committee era, so they are shown here rather than in the all-years trends.")
            for field, title in [("tuition_growth", "Tuition growth assumption"),
                                 ("inv_return", "Investment return assumption"),
                                 ("benefit", "Benefit structure"),
                                 ("min_benefits", "Minimum benefits / alternative")]:
                rows = []
                for k in plan_order([kk for kk in d["plan_key"].unique()
                                     if kk not in ("MI-I",)]):  # fold Michigan to one row
                    dk = "MI" if k == "MI-II" else k
                    if dk == "MI" and k != "MI-II":
                        continue
                    disp_name = "Michigan Education Trust (MET)" if dk == "MI" else name_of(k)
                    rec = {"Plan": disp_name}
                    any_val = False
                    for y in era_years:
                        akey = "MI-II" if dk == "MI" else dk
                        val = (aby.get(str(y), {}).get(akey, {}) or {}).get(field, "")
                        if "unavailable" in str(val).lower():
                            val = ""
                        if field in ("tuition_growth", "inv_return"):
                            val = _fmt_assumption(val)
                        rec[str(y)] = val
                        any_val = any_val or bool(val)
                    if any_val:
                        rows.append(rec)
                st.markdown(f"**{title}**")
                st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True)

        else:  # Metric matrix
            metric = st.selectbox("Metric", NUM_COLS, format_func=lambda m: METRICS[m][0], key="eram")
            de = d[d["reporting_year"].isin(era_years)]
            look = {(r["plan_key"], int(r["reporting_year"])): r[metric] for _, r in de.iterrows()}
            rows = []
            for k in plan_order(de["plan_key"].unique()):
                rec = {"Plan": name_of(k), "Status": REG.get(k, ["","","","",0])[3]}
                for y in era_years:
                    v = look.get((k, y))
                    rec[str(y)] = fmt(v, METRICS[metric][1]) if v is not None and not pd.isna(v) else ""
                rows.append(rec)
            st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True, height=560)

# =========================================================== DATA QUALITY
with tabs[5]:
    d = df()
    years = sorted(d["reporting_year"].unique())
    keys = plan_order(d["plan_key"].unique())

    st.markdown("#### Completeness by metric")
    comp = []
    for m in NUM_COLS:
        filled = d[m].notna().sum()
        comp.append({"Metric": METRICS[m][0], "Filled": int(filled),
                     "Of rows": len(d), "Coverage": f"{filled/len(d)*100:.0f}%"})
    st.dataframe(pd.DataFrame(comp), width='stretch', hide_index=True)

    st.markdown("#### Flags to review")
    flags = []
    for _, r in d.iterrows():
        k, y, note, asof = r["plan_key"], int(r["reporting_year"]), str(r["note"]).lower(), str(r["as_of"])
        if r["active_accounts"] and pd.notna(r["active_accounts"]) and pd.isna(r["funded"]):
            flags.append((name_of(k), y, "Funded status blank though accounts reported"))
        if "unavailable" in note or "carried" in note:
            flags.append((name_of(k), y, "Carried-forward / updated data unavailable"))
        m = re.search(r"(20\d\d)", asof)
        if m and int(m.group(1)) < y:
            flags.append((name_of(k), y, f"Reported as of {asof} (earlier than collection year)"))
        if pd.notna(r["funded"]) and (r["funded"] > 4 or r["funded"] < 0.15):
            flags.append((name_of(k), y, f"Outlier funded status ({r['funded']*100:.0f}%) — confirm"))
        # scale sanity: a plan cannot have fewer lifetime accounts than currently active ones
        if pd.notna(r["accounts_since_inception"]) and pd.notna(r["active_accounts"]) \
           and r["accounts_since_inception"] < r["active_accounts"]:
            flags.append((name_of(k), y, "Accounts since inception is below active accounts — "
                                         "likely a unit or scale error in the source"))
    if flags:
        st.dataframe(pd.DataFrame(flags, columns=["Plan", "Year", "Flag"]),
                     width='stretch', hide_index=True)
    else:
        st.success("No flags.")

    st.markdown("#### Missing plan-years (within collected years)")
    miss = []
    for k in keys:
        pres = set(d[d["plan_key"] == k]["reporting_year"])
        # only count years after a plan's first appearance (avoids flagging pre-existence)
        if not pres: continue
        for y in years:
            if min(pres) <= y <= max(pres) and y not in pres:
                miss.append((name_of(k), y))
    st.caption(f"{len(miss)} interior gaps (a plan skipped a year between its first and last report).")
    if miss:
        st.dataframe(pd.DataFrame(miss, columns=["Plan", "Missing year"]),
                     width='stretch', hide_index=True)

# =========================================================== UPLOAD & APPEND
with tabs[6]:
    st.markdown("#### Add a new year")
    st.caption("Upload a standard Prepaid Committee Excel file. The parser reads columns by "
               "their headers, so both the older (2022–23) and newer (2024–25) layouts work. "
               "Michigan MET I/MET II and Mississippi tiers are split automatically.")

    tc = st.columns([1, 1, 2])
    up_year = tc[0].number_input("Reporting year for this file", min_value=2000, max_value=2100,
                                 value=datetime.now().year, step=1)
    st.download_button("Download blank template", template_bytes(),
                       file_name="Prepaid_Committee_Template.xlsx", key="tmpl")

    upfile = st.file_uploader("Committee Excel (.xlsx)", type=["xlsx"])
    if upfile is not None:
        try:
            recs, warns, up_attrs = parse_committee_file(upfile, int(up_year))
            new = enrich(recs, int(up_year), f"Committee upload (FY{int(up_year)})")
            st.success(f"Parsed {len(new)} plan rows for {int(up_year)}.")
            for w in warns: st.warning(w)

            existing_keys = set(df()[df()["reporting_year"] == int(up_year)]["plan_key"])
            conflicts = [name_of(k) for k in new["plan_key"] if k in existing_keys]
            prev = new[["state","plan_name","funded","assets_m","active_accounts",
                        "accounts_since_inception","paid_out_fy_m","paid_out_inception_m","as_of","note"]].copy()
            prev["funded"] = prev["funded"].apply(lambda v: fmt(v, "pct"))
            st.dataframe(prev, width='stretch', hide_index=True)

            if conflicts:
                st.warning(f"{int(up_year)} already has data for: {', '.join(conflicts)}.")
            mode = st.radio("If the year already exists:",
                            ["Overwrite existing rows for this year", "Skip plans already present"],
                            horizontal=True)
            if st.button("Merge into session", type="primary"):
                base = df().copy()
                if mode.startswith("Overwrite"):
                    base = base[~((base["reporting_year"] == int(up_year)) &
                                  (base["plan_key"].isin(new["plan_key"])))]
                    add = new
                else:
                    add = new[~new["plan_key"].isin(existing_keys)]
                st.session_state.df = pd.concat([base, add], ignore_index=True)
                st.session_state.attrs_by_year[str(int(up_year))] = up_attrs
                st.success(f"Merged. The record now holds {st.session_state.df['reporting_year'].nunique()} years "
                           f"and {len(st.session_state.df)} rows. Download the new master below to commit it.")
        except Exception as e:
            st.error(f"Could not parse this file: {e}")

    st.divider()
    st.markdown("#### Save the new year permanently")
    st.info("Two files make up the record. Download **both** and commit them, or the new year's "
            "actuarial assumptions and descriptive columns will be lost on the next reboot.", icon="💾")
    dl = st.columns(2)
    dl[0].download_button("1. prepaid_master.csv  (the figures)",
                          df().to_csv(index=False).encode(),
                          file_name="prepaid_master.csv", type="primary")
    dl[1].download_button("2. prepaid_attrs_by_year.json  (assumptions + attributes)",
                          json.dumps(st.session_state.attrs_by_year, ensure_ascii=False, indent=0).encode(),
                          file_name="prepaid_attrs_by_year.json", type="primary")
    st.caption("Replace `data/prepaid_master.csv` and `data/prepaid_attrs_by_year.json` in the repo, "
               "then commit and push. Streamlit Cloud redeploys automatically.")

    st.divider()
    st.markdown("#### Other downloads")
    st.download_button("Consolidated analytical workbook (Excel)", build_workbook(df()),
                       file_name=f"CSPN_Prepaid_Consolidated_{min(df()['reporting_year'])}-{max(df()['reporting_year'])}.xlsx")

# =========================================================== EXPORTS
with tabs[7]:
    d = df()
    years = sorted(d["reporting_year"].unique())
    aby = st.session_state.attrs_by_year

    st.markdown("#### Committee-format one-pager (Excel)")
    st.caption("Reproduces the committee sheet layout. Pick one year for a single formatted sheet, "
               "or several years to get them combined into one workbook (a tab per year). "
               "Michigan MET I/II and Mississippi tiers are recombined into the committee's own cell style.")
    sel_years = st.multiselect("Year(s)", list(reversed(years)), default=[max(years)])
    if sel_years:
        sel_years = sorted(sel_years)
        preview_year = st.selectbox("Preview year", list(reversed(sel_years)))
        op_rows, cl_rows = committee_rows(d, preview_year, aby.get(str(preview_year), {}), aby)
        prev = pd.DataFrame(
            [{**{"Section": "Open"}, **{k: v for k, v in r.items() if k != "_funded_num"}} for r in op_rows] +
            [{**{"Section": "Closed"}, **{k: v for k, v in r.items() if k != "_funded_num"}} for r in cl_rows])
        st.dataframe(prev, width='stretch', hide_index=True)

        fname = (f"Prepaid_Committee_{sel_years[0]}.xlsx" if len(sel_years) == 1
                 else f"Prepaid_Committee_{sel_years[0]}-{sel_years[-1]}.xlsx")
        label = ("Download committee-format Excel" if len(sel_years) == 1
                 else f"Download {len(sel_years)}-year committee workbook")
        st.download_button(label, committee_workbook(d, sel_years, aby),
                           file_name=fname, type="primary")
    else:
        st.info("Select at least one year.")

    st.divider()
    st.markdown("#### Full analytical master (Excel)")
    st.caption("The consolidated workbook: README, long table, and Funded / Assets / Accounts matrices across every year.")
    st.download_button("Download analytical master", build_workbook(d),
                       file_name=f"CSPN_Prepaid_Consolidated_{min(years)}-{max(years)}.xlsx")

    st.divider()
    st.markdown("#### Filtered raw data (CSV)")
    fc = st.columns(3)
    yy = fc[0].multiselect("Years ", years, default=years, key="exy")
    stt = fc[1].multiselect("Status ", sorted(d["status"].unique()), default=sorted(d["status"].unique()), key="exs")
    kk = fc[2].multiselect("Plans ", plan_order(d["plan_key"].unique()),
                           default=plan_order(d["plan_key"].unique()), format_func=name_of, key="exk")
    filt = d[d["reporting_year"].isin(yy) & d["status"].isin(stt) & d["plan_key"].isin(kk)]
    st.caption(f"{len(filt)} rows selected.")
    st.download_button("Download filtered CSV", filt.to_csv(index=False).encode(),
                       file_name="prepaid_filtered.csv")
