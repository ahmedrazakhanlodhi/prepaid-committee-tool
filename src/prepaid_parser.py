"""
prepaid_parser.py
Header-driven parser for CSPN / The 529 Network annual Prepaid Committee Excel files.
Handles the 2022/2023 layout and the 2024/2025 layout uniformly by locating columns
from their header labels rather than fixed positions. Returns tidy long-format records.
"""
import re
import openpyxl

# canonical plan registry key resolution
STATEMAP = {"AK":"AK","FL":"FL","MA":"MA","MD":"MD","MS":"MS","NV":"NV","PA":"PA","VA":"VA",
            "WA":"WA","AL":"AL","IL":"IL","KY":"KY","OH":"OH","SC":"SC","CO":"CO","TN":"TN","WV":"WV"}

FIELD_KEYS = {
    "fund":   ["funded status"],
    "asof":   ["as of date", "as of"],
    "acct":   ["active accounts"],
    "assets": ["current assets"],
    "asi":    ["accounts since inception"],
    "poy":    ["paid out in most recent", "most recent fiscal", "paid out this"],
    "poi":    ["paid out since inception"],
    "notes":  ["additional notes"],
    "tuition":["tuition growth"],
    "invret": ["investment return"],
    "ff":     ["full faith"],
    "tax":    ["state tax"],
    "benefit":["benefit structure"],
    "estab":  ["year established"],
    "minben": ["minimum benefit", "alternative"],
}

def _p_funded(raw):
    if raw is None: return None
    s = str(raw).strip().replace(",", "")
    if s == "" or s.upper() in ("N/A","NA","-","OPEN","CLOSED","UNAVAILABLE"): return None
    m = re.search(r'([\d\.]+)\s*%', s)
    if m: return round(float(m.group(1))/100, 4)
    try: return round(float(s.rstrip("*").strip()), 4)
    except: return None

def _p_money_m(raw):
    if raw is None: return None
    s = str(raw).strip().replace("\n"," ").replace(",", "")
    if s == "" or s.upper() in ("N/A","NA","-","UNAVAILABLE"): return None
    s = s.replace("$","").strip()
    m = re.match(r'^([\d\.]+)\s*([BM]?)', s, re.I)
    if not m: return None
    v = float(m.group(1))
    if m.group(2).upper() == "B": v *= 1000.0
    return round(v, 3)

def _p_int(raw):
    if raw is None: return None
    s = str(raw).strip().replace("\n"," ").replace(",", "")
    if s.upper() in ("N/A","NA","-","UNAVAILABLE",""): return None
    m = re.match(r'^(\d+)', s)
    return int(m.group(1)) if m else None

def _norm_asof(v):
    if v is None: return ""
    if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
    return str(v).strip()

def _split_met(text):
    """Split a Michigan cell into (MET I, MET II) using B&C/D or MET I/MET II labels."""
    t = str(text)
    if re.search(r'MET\s*II', t):
        parts = re.split(r'MET\s*II\s*:?', t, maxsplit=1)
        m1 = re.sub(r'MET\s*I\s*:?', '', parts[0]).strip(" :")
        m2 = parts[1].strip(" :") if len(parts) > 1 else ""
        return m1, m2
    if re.search(r'\bB&C\b', t):
        parts = re.split(r'\s+D\s+', t, maxsplit=1)
        m1 = re.sub(r'B&C', '', parts[0]).strip(" :")
        m2 = parts[1].strip(" :") if len(parts) > 1 else ""
        return m1, m2
    return None, None

def _resolve_columns(ws, header_rows=range(3, 6)):
    """Scan candidate header rows and map each field to a column index by keyword."""
    colmap = {}
    max_c = ws.max_column
    for field, keys in FIELD_KEYS.items():
        found = None
        for r in header_rows:
            for c in range(1, max_c + 1):
                val = ws.cell(r, c).value
                if val is None: continue
                low = str(val).strip().lower()
                if any(k in low for k in keys):
                    found = c
                    break
            if found: break
        colmap[field] = found
    return colmap

def parse_committee_file(path_or_buffer, reporting_year):
    """
    Parse one committee workbook into long records.
    reporting_year: int the user assigns to this collection (e.g. 2026).
    Returns (records, warnings).
    """
    wb = openpyxl.load_workbook(path_or_buffer, data_only=True)
    ws = wb[wb.sheetnames[0]]
    C = _resolve_columns(ws)
    warnings = []
    required = ["fund", "acct", "assets"]
    missing_cols = [k for k in required if C.get(k) is None]
    if missing_cols:
        raise ValueError(
            "Could not find expected columns: " + ", ".join(missing_cols) +
            ". Make sure this is a standard Prepaid Committee file (with Funded Status, "
            "Active Accounts, Current Assets headers)."
        )

    records = []
    attrs = {}
    section = None
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None: continue
        a = str(a).strip()
        up = a.upper()
        if up.startswith("OPEN PLANS"):   section = "open";   continue
        if up.startswith("CLOSED PLANS"): section = "closed"; continue
        if a[:1] in ("¹","²","*","³") or "Alaska," in a:      continue

        def cell(field):
            ci = C.get(field)
            return ws.cell(r, ci).value if ci else None

        def clean(field):
            v = cell(field)
            return str(v).replace("\n", " ").strip() if v is not None else ""

        row_attr = dict(
            established=clean("estab"), full_faith=clean("ff"), tax=clean("tax"),
            benefit=clean("benefit"), min_benefits=clean("minben"),
            tuition_growth=clean("tuition"), inv_return=clean("invret"),
        )
        note = str(cell("notes") or "").replace("\n", " ").strip()
        asof = _norm_asof(cell("asof"))

        if a == "MI":
            attrs["MI-I"] = row_attr; attrs["MI-II"] = dict(row_attr)
            f1, f2 = _split_met(cell("fund"))
            ac1, ac2 = _split_met(cell("acct"))
            as1, as2 = _split_met(cell("assets"))
            si1, si2 = _split_met(cell("asi"))
            py1, py2 = _split_met(cell("poy"))
            records.append(dict(plan_key="MI-I", reporting_year=reporting_year, as_of=asof,
                funded=_p_funded(f1), assets_m=_p_money_m(as1), active_accounts=_p_int(ac1),
                accounts_since_inception=_p_int(si1), paid_out_fy_m=_p_money_m(py1),
                paid_out_inception_m=None,
                note=("MET I (Plans B & C)" + ("; " + note if note else "")).strip()))
            records.append(dict(plan_key="MI-II", reporting_year=reporting_year, as_of=asof,
                funded=_p_funded(f2), assets_m=_p_money_m(as2), active_accounts=_p_int(ac2),
                accounts_since_inception=_p_int(si2), paid_out_fy_m=_p_money_m(py2),
                paid_out_inception_m=None,
                note=("MET II (Plan D)" + ("; " + note if note else "")).strip()))
            continue

        if a == "TX":
            key = "TX-II" if section == "open" else "TX-I"
        elif a == "U.S.":
            key = "US"
        elif a in STATEMAP:
            key = STATEMAP[a]
        else:
            if re.match(r'^[A-Za-z\.]{2,4}$', a):
                warnings.append(f"Unrecognized plan code '{a}' on row {r} was skipped.")
            continue

        attrs[key] = row_attr

        fund_txt = str(cell("fund") or "")
        if "Horizon" in fund_txt or "Legacy" in fund_txt:
            h = re.search(r'Horizon\s*:?\s*([\d\.]+)\s*%', fund_txt)
            l = re.search(r'Legacy\s*:?\s*([\d\.]+)\s*%', fund_txt)
            funded = round(float(h.group(1))/100, 4) if h else None
            if l: note = (f"Legacy tier {l.group(1)}%" + ("; " + note if note else ""))
        else:
            funded = _p_funded(fund_txt)

        records.append(dict(plan_key=key, reporting_year=reporting_year, as_of=asof,
            funded=funded, assets_m=_p_money_m(cell("assets")),
            active_accounts=_p_int(cell("acct")),
            accounts_since_inception=_p_int(cell("asi")),
            paid_out_fy_m=_p_money_m(cell("poy")),
            paid_out_inception_m=_p_money_m(cell("poi")), note=note))

    return records, warnings, attrs
